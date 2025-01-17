#Corpuz, Mark Jhay
#Deang, April Joy
#Espero, Airysh Xander

import customtkinter
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

#Setting up main window
root = customtkinter.CTk()
root.title("Finance Tracker")
root.geometry("410x650")
root.maxsize(410, 650)
root.minsize(410, 650)
root.configure(fg_color=("#61876E"))

#App Class
class FinanceTracker():

    def __init__(self):
        super().__init__()

        # --- Upper Title Frame -----------------------------------------------------------------------
        self.BorderFrame = customtkinter.CTkFrame(root, width=410, height=40, corner_radius=0, fg_color=("#3C6255"))
        self.BorderFrame.pack()

        # Upper Title Frame Label
        self.BorderFrameLabel = customtkinter.CTkLabel(self.BorderFrame, text="EXPENSE TRACKER", font=("Arial Black", 18), text_color=("white"), bg_color=("#3C6255"))
        self.BorderFrameLabel.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # --- Upper Frame -----------------------------------------------------------------------
        self.UpperFrame = customtkinter.CTkFrame(root, width=340, height=280, corner_radius=20, fg_color=("#3C6255"), bg_color=("#61876E"))
        self.UpperFrame.pack(padx=10, pady=10)

        # Upper Frame Label
        self.UpperFrameLabel = customtkinter.CTkLabel(root, text="Your Financial Graph", font=("Arial Black", 18), text_color=("white"), bg_color=("#3C6255"))
        self.UpperFrameLabel.place(relx=0.5, rely=0.1, anchor=tk.CENTER)

        # --- Lower Frame ------------------------------------------------------------------------------
        self.LowerFrame = customtkinter.CTkFrame(root, width=340, height=280, corner_radius=20, fg_color=("#3C6255"))
        self.LowerFrame.pack(padx=10, pady=10)

        # Opening of an excel file and specific worksheet number, then storing an index value to balance
        wb = load_workbook("Expenses.xlsx")
        ws = wb["Sheet3"]
        balance = int(ws["A2"].value)

        # Asking for User's balance if balance is 0
        if balance == 0:

            # Asking for Balance Label
            self.AskBalanceLabel = tk.Label(self.LowerFrame, text="ADD YOUR BUDGET FIRST!", font=('Arial Black', 8), fg="white", bg="#3C6255")
            self.AskBalanceLabel.grid(column=1, row=0, pady=(10,30))

            # Enter Balance Label
            self.EnterBalanceLabel = tk.Label(self.LowerFrame, text="Enter Initial Budget:", fg="white", bg="#3C6255")
            self.EnterBalanceLabel.grid(column=1, row=1, padx=20)

            # Entry for balance input
            self.BalanceEntry = tk.Entry(self.LowerFrame, width=20)
            self.BalanceEntry.grid(column=1, row=2, padx=(0,15), pady=(0,50))

            # Save button 
            def SaveButton():
                
                # Storing the user balance input to an index in excel file
                ws["A2"].value = int(self.BalanceEntry.get())
                wb.save("Expenses.xlsx") # Saving the excel file

                # Destroying widgets
                self.AskBalanceLabel.destroy() 
                self.EnterBalanceLabel.destroy()
                self.BalanceEntry.destroy()
                self.ButtonSave.destroy()
                
                # Initializing Mainmenu in lower frame
                self.Mainmenu()

            # Widget button for save
            self.ButtonSave = tk.Button(self.LowerFrame, text='SAVE', command=SaveButton)
            self.ButtonSave.grid(column=1, row=5, pady=(30,10))

        else:
            # Initializing Mainmenu in lower frame
            self.Mainmenu()
        
    # Class method for Pie Chart
    def Graph(self):
        
        try:
            # Opening excel file
            wb = load_workbook("Expenses.xlsx")
            ws = wb["Sheet3"]
            ws2 = wb["Sheet3"]

            # Getting the index values and storing it in TotalEpenses and TotalProfits
            TotalExpenses = int(ws["B2"].value)
            TotalBalance = int(ws2["A2"].value)
            wb.save("Expenses.xlsx") # Saving excel file

            # Setting figure size and pie size of the Pie Chart
            self.fig = plt.figure(figsize=(5,5), dpi=100)
            self.fig.set_size_inches(3.7, 2.7)
            self.fig.set_facecolor('#3C6255') # Setting its background color to match with its background frame

            PieChart = [TotalExpenses, TotalBalance] # Storing values
            PieLabel = ["Total Expenses", "Total Budget"] # Storing labels
            
            # Enhanced color palette
            colors = ['#FF6F61', '#6CA0DC']

            # Configuring the Pie Chart
            wedges, texts, autotexts = plt.pie(
                PieChart,
                radius=0.7,
                autopct="%.1f%%",
                shadow=True,
                explode=[0, 0.1],
                colors=colors,
                startangle=140
            )

            # Adding a legend to display labels outside the pie chart
            plt.legend(
                wedges, 
                PieLabel, 
                loc="lower left", 
                bbox_to_anchor=(0.8, 0.3), 
                fontsize=10, 
                frameon=False
            )

            # Customizing percentage labels
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontsize(15)
                autotext.set_fontweight('bold')

            # Adjusting layout for better spacing
            self.fig.tight_layout(rect=[0, 0, 0.9, 1])  # Adjusted space for the legend

            # Setting canvas for Pie Chart 
            self.canvasbar = FigureCanvasTkAgg(self.fig, self.UpperFrame)
            self.canvasbar.draw() # Drawing the canvas
            self.canvasbar.get_tk_widget().pack(anchor=tk.CENTER) # Placing the Pie Chart and Canvas
        
        except RuntimeError:
            pass


    # Method for Mainmenu in lower frame
    def Mainmenu(self):
        
        # Button widget
        self.button1 = tk.Button(self.LowerFrame, text="+ Add Expense", font=("Arial Black", 8), fg="white", bg="#61876E", command=self.AddExpense, activebackground="#0081C9")
        self.button1.grid(column=0, row=0, padx=15, pady=10), 

        # Button widget
        self.button2 = tk.Button(self.LowerFrame, text="+ Add Balance", font=("Arial Black", 8), fg="white", bg="#61876E", command=self.AddProfit, activebackground="#0081C9")
        self.button2.grid(column=1, row=0, padx=15, pady=10)

        # Button widget
        self.button3 = tk.Button(self.LowerFrame, text="Check History", font=("Arial Black", 8), fg="white", bg="#61876E", command=self.CheckHistory, activebackground="#0081C9")
        self.button3.grid(column=2, row=0, padx=15, pady=10)

        # Label widget
        self.label1 = tk.Label(self.LowerFrame, text="RECENT EXPENSES:", font=("Arial Black", 12), fg="white", bg="#3C6255")
        self.label1.grid(columnspan=3, row=1, pady=(40,2), ipadx=100)

        # To read and store the whole index table
        df = pd.read_excel('Expenses.xlsx')

        # Create a Treeview for a clean table design
        self.expenses_table = ttk.Treeview(self.LowerFrame, columns=("Product Name", "Product Type", "Product Cost", "Date of Purchase"), show="headings", height=3)

        # Define column headings
        self.expenses_table.heading("Product Name", text="Product Name")
        self.expenses_table.heading("Product Type", text="Product Type")
        self.expenses_table.heading("Product Cost", text="Product Cost")
        self.expenses_table.heading("Date of Purchase", text="Date of Purchase")

        # Set column widths
        self.expenses_table.column("Product Name", width=80, anchor=tk.W)
        self.expenses_table.column("Product Type", width=80, anchor=tk.W)
        self.expenses_table.column("Product Cost", width=80, anchor=tk.W)
        self.expenses_table.column("Date of Purchase", width=80, anchor=tk.W)

        # Add some padding to the Treeview
        self.expenses_table.grid(columnspan=3, row=2, pady=(0, 15), ipadx=30)

        # Populate the table with the last 3 rows from the DataFrame
        recent_expenses = df.tail(3).values.tolist()  # Convert last 3 rows to a list
        for expense in recent_expenses:
            self.expenses_table.insert("", "end", values=expense)
        
        # Opening the excel file
        wb = load_workbook("Expenses.xlsx")
        ws = wb["Sheet3"] # Setting sheet 3 as the active sheet
        NewBalance = int(ws["A2"].value) # Storing the index value
        
        # Label widget
        self.label2 = tk.Label(self.LowerFrame, text="REMAINING BALANCE: ", font=("Arial Black", 12), fg="white", bg="#3C6255")
        self.label2.grid(columnspan=2, row=3, pady=(25,15), ipadx=20)

        # Label widget for printing the user balance
        self.label3 = tk.Label(self.LowerFrame, text=NewBalance, font=("Arial Black", 12), fg="white", bg="#3C6255")
        self.label3.grid(column=2, row=3, pady=(25,15), ipadx=25)

        # Initializing pie chart
        self.Graph()

    
    # Add Expense
    def AddExpense(self):
        
        # Submit Button
        def ButtonSubmit():

            # Opening the excel file
            wb = load_workbook("Expenses.xlsx")
            ws = wb["Sheet3"] # Setting sheet 3 as the active sheet

            # Subtracting the expenses from the total balance and storing its new value
            NewBalance = (int(ws["A2"].value) - int(self.ProductCostEntry.get()))
            ws["A2"].value = NewBalance

            # Adding the expenses to the Total Expense
            TotalExpense = (int(ws["B2"].value) + int(self.ProductCostEntry.get()))
            ws["B2"].value = TotalExpense
            wb.save("Expenses.xlsx") # Saving the excel file
            
            # Opening the excel file
            wb = load_workbook("Expenses.xlsx")
            ws = wb["Sheet1"] # Setting sheet 1 as the active sheet

            # Appending the purchased product into the excel sheet 1
            ws.append([self.ProductTypeCombobox.get(), self.ProductNameEntry.get(), self.ProductCostEntry.get(), self.DateOfPurchaseEntry.get()])
            wb.save("Expenses.xlsx") # Saving the excel file
        
            self.ErrorMessage = messagebox.showerror("Error!", "Insufficient funds!")

            blank = "" # To be used for emptying the entry boxes

            # To clear entry boxes after clicking submit button
            self.ProductTypeCombobox.delete("0", "end")
            self.ProductNameEntry.delete("0", "end")
            self.ProductNameEntry.insert("0", blank)
            self.ProductCostEntry.delete("0", "end")
            self.ProductCostEntry.insert("0", blank)
            self.DateOfPurchaseEntry.delete("0", "end")
            self.DateOfPurchaseEntry.insert("0", blank)
            self.canvasbar.get_tk_widget().destroy()
           
            self.Graph() # Initializing Pie Chart

        # Back button
        def ButtonBack():

            # Destroying previous widgets to be changed with a new one
            self.AddExpenseLabel.destroy()
            self.ProductTypeLabel.destroy()
            self.ProductTypeCombobox.destroy()
            self.ProductNameLabel.destroy()
            self.ProductNameEntry.destroy()
            self.ProductCostLabel.destroy()
            self.ProductCostEntry.destroy()
            self.DateOfPurchaseLabel.destroy()
            self.DateOfPurchaseEntry.destroy()
            self.SubmitButton.destroy()
            self.BackButton.destroy()
            self.canvasbar.get_tk_widget().destroy()
            
            # Button widget
            self.button1 = tk.Button(self.LowerFrame, text="+ Add Expense", font=("Arial Black", 8), fg="white", bg="#61876E", command=self.AddExpense, activebackground="#0081C9")
            self.button1.grid(column=0, row=0, padx=15, pady=10), 

            # Button widget
            self.button2 = tk.Button(self.LowerFrame, text="+ Add Balance", font=("Arial Black", 8), fg="white", bg="#61876E", command=self.AddProfit, activebackground="#0081C9")
            self.button2.grid(column=1, row=0, padx=15, pady=10)

            # Button widget
            self.button3 = tk.Button(self.LowerFrame, text="Check History", font=("Arial Black", 8), fg="white", bg="#61876E", command=self.CheckHistory, activebackground="#0081C9")
            self.button3.grid(column=2, row=0, padx=15, pady=10)

            # Label widget
            self.label1 = tk.Label(self.LowerFrame, text="RECENT EXPENSES:", font=("Arial Black", 12), fg="white", bg="#3C6255")
            self.label1.grid(columnspan=3, row=1, pady=(40,2), ipadx=100)

            # To read and store the existing values from the table
            df = pd.read_excel('Expenses.xlsx')

            # Create a Treeview for a clean table design
            self.expenses_table = ttk.Treeview(self.LowerFrame, columns=("Product Name", "Product Type", "Product Cost", "Date of Purchase"), show="headings", height=3)

            # Define column headings
            self.expenses_table.heading("Product Name", text="Product Name")
            self.expenses_table.heading("Product Type", text="Product Type")
            self.expenses_table.heading("Product Cost", text="Product Cost")
            self.expenses_table.heading("Date of Purchase", text="Date of Purchase")

            # Set column widths
            self.expenses_table.column("Product Name", width=80, anchor=tk.W)
            self.expenses_table.column("Product Type", width=80, anchor=tk.W)
            self.expenses_table.column("Product Cost", width=80, anchor=tk.W)
            self.expenses_table.column("Date of Purchase", width=80, anchor=tk.W)

            # Add some padding to the Treeview
            self.expenses_table.grid(columnspan=3, row=2, pady=(0, 15), ipadx=30)

            # Populate the table with the last 3 rows from the DataFrame
            recent_expenses = df.tail(3).values.tolist()  # Convert last 3 rows to a list
            for expense in recent_expenses:
                self.expenses_table.insert("", "end", values=expense)
            
            # Opening the excel file
            wb = load_workbook("Expenses.xlsx")
            ws = wb["Sheet3"] # Setting sheet 3 as the active sheet
            NewBalance = int(ws["A2"].value) # Storing the index value
            
            # Label widget
            self.label2 = tk.Label(self.LowerFrame, text="REMAINING BALANCE: ", font=("Arial Black", 12), fg="white", bg="#3C6255")
            self.label2.grid(columnspan=2, row=3, pady=(25,15), ipadx=20)

            # Label widget for printing the user balance
            self.label3 = tk.Label(self.LowerFrame, text=NewBalance, font=("Arial Black", 12), fg="white", bg="#3C6255")
            self.label3.grid(column=2, row=3, pady=(25,15), ipadx=25)

            # Initializing pie chart
            self.Graph()

        # Destroying previous widgets to be changed with a new one
        self.expenses_table.destroy()
        self.button1.destroy()
        self.button2.destroy()
        self.button3.destroy()
        self.label1.destroy()
        self.label2.destroy()
        self.label3.destroy()

        # Label widget
        self.AddExpenseLabel = tk.Label(self.LowerFrame, text="ADD EXPENSE", font=('Arial Black', 8), fg="white", bg="#3C6255")
        self.AddExpenseLabel.grid(column=1, row=0, pady=10)
        
        # Label widget
        self.ProductTypeLabel = tk.Label(self.LowerFrame, text="Select Product Type:", fg="white", bg="#3C6255")
        self.ProductTypeLabel.grid(column=0, row=1, padx=(15,0))

        # Combobox widget
        ProductTypeList = ["Food & Drinks", "Gadgets & Electronics", "Clothing", "House Utilities", "Entertainment"]
        self.ProductTypeCombobox = ttk.Combobox(self.LowerFrame, value=ProductTypeList, width=18)
        self.ProductTypeCombobox.grid(column=0, row=2, padx=(15,0), pady=(0, 25))
        
        # Label widget
        self.ProductNameLabel = tk.Label(self.LowerFrame, text="Product Name:", fg="white", bg="#3C6255")
        self.ProductNameLabel.grid(column=2, row=1, padx=(0,15))

        # Entry box widget
        self.ProductNameEntry = tk.Entry(self.LowerFrame, width=20)
        self.ProductNameEntry.grid(column=2, row=2, padx=(0,15), pady=(0, 25))

        # Label widget
        self.ProductCostLabel = tk.Label(self.LowerFrame, text="Product Cost:", fg="white", bg="#3C6255")
        self.ProductCostLabel.grid(column=0, row=3, padx=(15,0))

        # Entry box widget
        self.ProductCostEntry = tk.Entry(self.LowerFrame, width=20)
        self.ProductCostEntry.grid(column=0, row=4, padx=(15,0), pady=(0, 25))

        # Label widget
        self.DateOfPurchaseLabel = tk.Label(self.LowerFrame, text="Date of Purchase:", fg="white", bg="#3C6255")
        self.DateOfPurchaseLabel.grid(column=2, row=3, padx=(0,15))

        # Entry box widget
        self.DateOfPurchaseEntry = tk.Entry(self.LowerFrame, width=20)
        self.DateOfPurchaseEntry.grid(column=2, row=4, padx=(0,15), pady=(0, 25))

        # Button widget
        self.SubmitButton = tk.Button(self.LowerFrame, text='SUBMIT', command=ButtonSubmit, fg="white", bg=("#61876E"))
        self.SubmitButton.grid(column=0, row=5, padx=(15,0), pady=(30,10))

        # Button widget
        self.BackButton = tk.Button(self.LowerFrame, text='BACK', command=ButtonBack)
        self.BackButton.grid(column=2, row=5, padx=(0,15), pady=(30,10))

    # Add Profits
    def AddProfit(self):
        
        # Submit Button
        def ButtonSubmit():

            # Opening the excel file
            wb = load_workbook("Expenses.xlsx")
            ws = wb["Sheet3"] # Setting sheet 3 as the active sheet

            # Adding the profits to the total balance and storing its new value
            NewBalance = (int(ws["A2"].value) + int(self.ProfitAmountEntry.get()))
            ws["A2"].value = NewBalance 

            wb.save("Expenses.xlsx")
            
            blank = "" # To be used for emptying the entry boxes

            # To clear entry boxes after clicking submit button
            self.ProfitAmountEntry.delete("0", "end")
            self.ProfitAmountEntry.insert("0", blank)
            self.canvasbar.get_tk_widget().destroy()
            
            # Initializing Pie Chart
            self.Graph()

        # Back button
        def ButtonBack():

            # Destroying previous widgets to be changed with a new one
            self.expenses_table
            self.AddProfitLabel.destroy()
            self.ProfitAmountLabel.destroy()
            self.ProfitAmountEntry.destroy()
            self.SubmitButton.destroy()
            self.BackButton.destroy()
            self.canvasbar.get_tk_widget().destroy()
            
            # Button widget
            self.button1 = tk.Button(self.LowerFrame, text="+ Add Expense", font=("Arial Black", 8), fg="white", bg="#61876E", command=self.AddExpense, activebackground="#0081C9")
            self.button1.grid(column=0, row=0, padx=15, pady=10), 

            # Button widget
            self.button2 = tk.Button(self.LowerFrame, text="+ Add Balance", font=("Arial Black", 8), fg="white", bg="#61876E", command=self.AddProfit, activebackground="#0081C9")
            self.button2.grid(column=1, row=0, padx=15, pady=10)

            # Button widget
            self.button3 = tk.Button(self.LowerFrame, text="Check History", font=("Arial Black", 8), fg="white", bg="#61876E", command=self.CheckHistory, activebackground="#0081C9")
            self.button3.grid(column=2, row=0, padx=15, pady=10)

            # Label widget
            self.label1 = tk.Label(self.LowerFrame, text="RECENT EXPENSES:", font=("Arial Black", 12), fg="white", bg="#3C6255")
            self.label1.grid(columnspan=3, row=1, pady=(40,2), ipadx=100)

            # To read and store the existing values from the table
            df = pd.read_excel('Expenses.xlsx')

            # Create a Treeview for a clean table design
            self.expenses_table = ttk.Treeview(self.LowerFrame, columns=("Product Name", "Product Type", "Product Cost", "Date of Purchase"), show="headings", height=3)

            # Define column headings
            self.expenses_table.heading("Product Name", text="Product Name")
            self.expenses_table.heading("Product Type", text="Product Type")
            self.expenses_table.heading("Product Cost", text="Product Cost")
            self.expenses_table.heading("Date of Purchase", text="Date of Purchase")

            # Set column widths
            self.expenses_table.column("Product Name", width=80, anchor=tk.W)
            self.expenses_table.column("Product Type", width=80, anchor=tk.W)
            self.expenses_table.column("Product Cost", width=80, anchor=tk.W)
            self.expenses_table.column("Date of Purchase", width=80, anchor=tk.W)

            # Add some padding to the Treeview
            self.expenses_table.grid(columnspan=3, row=2, pady=(0, 15), ipadx=30)

            # Populate the table with the last 3 rows from the DataFrame
            recent_expenses = df.tail(3).values.tolist()  # Convert last 3 rows to a list
            for expense in recent_expenses:
                self.expenses_table.insert("", "end", values=expense)

            # Opening the excel file
            wb = load_workbook("Expenses.xlsx")
            ws = wb["Sheet3"] # Setting sheet 3 as the active sheet
            NewBalance = int(ws["A2"].value) # Storing the index value
            
            # Label widget
            self.label2 = tk.Label(self.LowerFrame, text="REMAINING BALANCE: ", font=("Arial Black", 12), fg="white", bg="#3C6255")
            self.label2.grid(columnspan=2, row=3, pady=(25,15), ipadx=20)

            # Label widget for printing the user balance
            self.label3 = tk.Label(self.LowerFrame, text=NewBalance, font=("Arial Black", 12), fg="white", bg="#3C6255")
            self.label3.grid(column=2, row=3, pady=(25,15), ipadx=25)

            # Initializing pie chart
            self.Graph() 

        # Destroying previous widgets to be changed with a new one
        self.expenses_table.destroy()
        self.button1.destroy()
        self.button2.destroy()
        self.button3.destroy()
        self.label1.destroy()
        self.label2.destroy()
        self.label3.destroy()

        # Label widget
        self.AddProfitLabel = tk.Label(self.LowerFrame, text="ADD BALANCE", font=('Arial Black', 8), fg="white", bg="#3C6255")
        self.AddProfitLabel.grid(column=1, row=0, pady=(10,60))

        # Label widget
        self.ProfitAmountLabel = tk.Label(self.LowerFrame, text="Enter Balance Amount:", fg="white", bg="#3C6255")
        self.ProfitAmountLabel.grid(column=1, row=1)

        # Entry box widget
        self.ProfitAmountEntry = tk.Entry(self.LowerFrame, width=18)
        self.ProfitAmountEntry.grid(column=1, row=2, padx=(15,0))

        # Button widget
        self.SubmitButton = tk.Button(self.LowerFrame, text='SUBMIT', command=ButtonSubmit, fg="white", bg=("#61876E"))
        self.SubmitButton.grid(column=0, row=3, padx=(15,0), pady=(30,10))

        # Button widget
        self.BackButton = tk.Button(self.LowerFrame, text='BACK', command=ButtonBack)
        self.BackButton.grid(column=2, row=3, padx=(0,15), pady=(30,10))

    # Check History
    def CheckHistory(self):
        
        # Back Button
        def ButtonBack():

            # Destroying previous widgets to be changed with a new one
            self.label4.destroy()
            self.expenses_table.destroy()
            self.BackButton.destroy()
            self.canvasbar.get_tk_widget().destroy()

            # Button widget
            self.button1 = tk.Button(self.LowerFrame, text="+ Add Expense", font=("Arial Black", 8), fg="white", bg="#61876E", command=self.AddExpense, activebackground="#0081C9")
            self.button1.grid(column=0, row=0, padx=15, pady=10), 

            # Button widget
            self.button2 = tk.Button(self.LowerFrame, text="+ Add Profit", font=("Arial Black", 8), fg="white", bg="#61876E", command=self.AddProfit, activebackground="#0081C9")
            self.button2.grid(column=1, row=0, padx=15, pady=10)

            # Button widget
            self.button3 = tk.Button(self.LowerFrame, text="Check History", font=("Arial Black", 8), fg="white", bg="#61876E", command=self.CheckHistory, activebackground="#0081C9")
            self.button3.grid(column=2, row=0, padx=15, pady=10)

            # Label widget
            self.label1 = tk.Label(self.LowerFrame, text="RECENT EXPENSES:", font=("Arial Black", 12), fg="white", bg="#3C6255")
            self.label1.grid(columnspan=3, row=1, pady=(40,2), ipadx=100)

            # To read and store the existing values from the table
            df = pd.read_excel('Expenses.xlsx')

            # Create a Treeview for a clean table design
            self.expenses_table = ttk.Treeview(self.LowerFrame, columns=("Product Name", "Product Type", "Product Cost", "Date of Purchase"), show="headings", height=3)

            # Define column headings
            self.expenses_table.heading("Product Name", text="Product Name")
            self.expenses_table.heading("Product Type", text="Product Type")
            self.expenses_table.heading("Product Cost", text="Product Cost")
            self.expenses_table.heading("Date of Purchase", text="Date of Purchase")

            # Set column widths
            self.expenses_table.column("Product Name", width=80, anchor=tk.W)
            self.expenses_table.column("Product Type", width=80, anchor=tk.W)
            self.expenses_table.column("Product Cost", width=80, anchor=tk.W)
            self.expenses_table.column("Date of Purchase", width=80, anchor=tk.W)

            # Add some padding to the Treeview
            self.expenses_table.grid(columnspan=3, row=2, pady=(0, 15), ipadx=30)

            # Populate the table with the last 3 rows from the DataFrame
            recent_expenses = df.tail(3).values.tolist()  # Convert last 3 rows to a list
            for expense in recent_expenses:
                self.expenses_table.insert("", "end", values=expense)

            # Opening the excel file
            wb = load_workbook("Expenses.xlsx")
            ws = wb["Sheet3"] # Setting sheet 3 as the active sheet
            NewBalance = int(ws["A2"].value) # Storing the index value
            
            # Label widget
            self.label2 = tk.Label(self.LowerFrame, text="REMAINING BALANCE: ", font=("Arial Black", 12), fg="white", bg="#3C6255")
            self.label2.grid(columnspan=2, row=3, pady=(25,15), ipadx=20)

            # Label widget
            self.label3 = tk.Label(self.LowerFrame, text=NewBalance, font=("Arial Black", 12), fg="white", bg="#3C6255")
            self.label3.grid(column=2, row=3, pady=(25,15), ipadx=25)

            # Initializing the pie chart
            self.Graph()

        # Destroying previous widgets to be changed with a new one
        self.expenses_table.destroy()
        self.button1.destroy()
        self.button2.destroy()
        self.button3.destroy()
        self.label1.destroy()
        self.label2.destroy()
        self.label3.destroy()

        # Label widget
        self.label4 = tk.Label(self.LowerFrame, text="EXPENSES HISTORY:", font=("Arial Black", 12), fg="white", bg="#3C6255")
        self.label4.grid(columnspan=3, row=1, pady=(40,2), ipadx=100)

        # To read and store the existing values from the table
        df = pd.read_excel('Expenses.xlsx')

        # Create a Treeview for a clean table design
        self.expenses_table = ttk.Treeview(self.LowerFrame, columns=("Product Name", "Product Type", "Product Cost", "Date of Purchase"), show="headings", height=5)

        # Define column headings
        self.expenses_table.heading("Product Name", text="Product Name")
        self.expenses_table.heading("Product Type", text="Product Type")
        self.expenses_table.heading("Product Cost", text="Product Cost")
        self.expenses_table.heading("Date of Purchase", text="Date of Purchase")

        # Set column widths
        self.expenses_table.column("Product Name", width=80, anchor=tk.W)
        self.expenses_table.column("Product Type", width=80, anchor=tk.W)
        self.expenses_table.column("Product Cost", width=80, anchor=tk.W)
        self.expenses_table.column("Date of Purchase", width=80, anchor=tk.W)

        # Add some padding to the Treeview
        self.expenses_table.grid(columnspan=3, row=2, pady=(0, 15), ipadx=30)

        # Populate the table with the last 3 rows from the DataFrame
        recent_expenses = df.tail(5).values.tolist()  # Convert last 3 rows to a list
        for expense in recent_expenses:
            self.expenses_table.insert("", "end", values=expense)

        # Button widget
        self.BackButton = tk.Button(self.LowerFrame, text='BACK', command=ButtonBack)
        self.BackButton.grid(column=1, row=4, pady=(30,10))
        
    # For starting the App
    def start(self):
        
        # To make sure that every process stops after clicking the window close button
        root.protocol("WM_DELETE_WINDOW", root.quit)
        root.mainloop()

#Starting the Class or the App
App = FinanceTracker()
App.start()